import pandas as pd
import io
from datetime import datetime
import pytz
from config import TIMEZONE

class ExcelHandler:
    def __init__(self):
        self.timezone = pytz.timezone(TIMEZONE)
    
    def create_purchase_request_template(self):
        """Создание шаблона заявки на покупку"""
        # Создаем DataFrame с заголовками и примерами
        data = {
            'Обект номи': ['Мисол: Жилой комплекс "Сам Сити"'],
            'Махсулот номи': ['Мисол: Цемент'],
            'Миқдори': ['100'],
            'Ўлчов бирлиги': ['мешок'],
            'Материал изох': ['Марка М400']
        }
        
        df = pd.DataFrame(data)
        
        # Создаем Excel файл в памяти
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Заявка')
            
            # Получаем рабочую книгу для форматирования
            workbook = writer.book
            worksheet = writer.sheets['Заявка']
            
            # Форматируем заголовки
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = cell.font.copy(bold=True)
                cell.fill = cell.fill.copy(fill_type='solid', fgColor='CCCCCC')
        
        output.seek(0)
        return output
    
    def create_seller_offer_template(self, request_data):
        """Создание шаблона предложения поставщика на основе заявки"""
        # Создаем DataFrame с товарами из заявки
        items = request_data.get('items', [])
        
        if not items:
            # Если нет товаров, создаем пустой шаблон
            data = {
                'Махсулот номи': [''],
                'Миқдори': [''],
                'Ўлчов бирлиги': [''],
                'Материал изох': [''],
                'нархи': [''],
                'Суммаси': ['']
            }
        else:
            # Создаем данные из товаров заявки
            data = {
                'Махсулот номи': [item['product_name'] for item in items],
                'Миқдори': [int(item['quantity']) for item in items],  # Преобразуем в int
                'Ўлчов бирлиги': [item['unit'] for item in items],
                'Материал изох': [item['material_description'] for item in items],
                'нархи': [''] * len(items),
                'Суммаси': [''] * len(items)
            }
        
        df = pd.DataFrame(data)
        
        # Создаем Excel файл в памяти
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Предложение')
            
            # Получаем рабочую книгу для форматирования
            workbook = writer.book
            worksheet = writer.sheets['Предложение']
            
            # Форматируем заголовки
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = cell.font.copy(bold=True)
                cell.fill = cell.fill.copy(fill_type='solid', fgColor='CCCCCC')
            
            # Выделяем поля для заполнения цен (желтые) и добавляем формулы
            for row in range(2, len(items) + 2):
                # Колонка нархи (желтая)
                price_cell = worksheet.cell(row=row, column=5)
                price_cell.fill = price_cell.fill.copy(fill_type='solid', fgColor='FFFF00')
                
                # Колонка Суммаси (желтая) с формулой
                total_cell = worksheet.cell(row=row, column=6)
                total_cell.fill = total_cell.fill.copy(fill_type='solid', fgColor='FFFF00')
                # Добавляем формулу =B{row}*E{row}
                total_cell.value = f"=B{row}*E{row}"
            
            # Устанавливаем формат для столбца Миқдори как целое число
            for row in range(2, len(items) + 2):
                quantity_cell = worksheet.cell(row=row, column=2)  # Столбец B (Миқдори)
                quantity_cell.number_format = '0'  # Формат целого числа
        
        output.seek(0)
        return output
    
    def parse_purchase_request(self, file_content):
        """Парсинг Excel файла с заявкой на покупку"""
        try:
            # Читаем Excel файл
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0)
            
            # Получаем общие данные заявки (из первой строки)
            first_row = df.iloc[0]
            object_name = str(first_row['Обект номи']).strip() if pd.notna(first_row['Обект номи']) and str(first_row['Обект номи']).strip() != '' else 'Не указан'
            
            # Убираем префикс "Мисол: " если он есть
            if object_name.startswith('Мисол: '):
                object_name = object_name[7:]
            
            items = []
            for index, row in df.iterrows():
                # Пропускаем пустые строки
                if pd.isna(row['Махсулот номи']) or row['Махсулот номи'] == '':
                    continue
                
                item = {
                    'product_name': str(row['Махсулот номи']) if pd.notna(row['Махсулот номи']) else 'Не указан',
                    'quantity': float(row['Миқдори']) if pd.notna(row['Миқдори']) else 0,
                    'unit': str(row['Ўлчов бирлиги']) if pd.notna(row['Ўлчов бирлиги']) else 'шт',
                    'material_description': str(row['Материал изох']) if pd.notna(row['Материал изох']) else ''
                }
                items.append(item)
            
            return {
                'object_name': object_name,
                'items': items
            }
            
        except Exception as e:
            raise Exception(f"Ошибка при парсинге Excel файла: {str(e)}")
    
    def parse_seller_offer(self, file_content):
        """Парсинг Excel файла с предложением поставщика"""
        try:
            # Читаем Excel файл
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0)
            
            items = []
            total_amount = 0
            
            for index, row in df.iterrows():
                # Пропускаем пустые строки
                if pd.isna(row['Махсулот номи']) or row['Махсулот номи'] == '':
                    continue
                
                # Проверяем, что заполнены цены
                if pd.isna(row['нархи']) or pd.isna(row['Суммаси']):
                    continue
                
                price_per_unit = float(row['нархи'])
                item_total = float(row['Суммаси'])
                total_amount += item_total
                
                item = {
                    'product_name': str(row['Махсулот номи']) if pd.notna(row['Махсулот номи']) else 'Не указан',
                    'quantity': float(row['Миқдори']) if pd.notna(row['Миқдори']) else 0,
                    'unit': str(row['Ўлчов бирлиги']) if pd.notna(row['Ўлчов бирлиги']) else 'шт',
                    'material_description': str(row['Материал изох']) if pd.notna(row['Материал изох']) else '',
                    'price_per_unit': price_per_unit,
                    'total_price': item_total
                }
                items.append(item)
            
            return {
                'items': items,
                'total_amount': total_amount
            }
            
        except Exception as e:
            raise Exception(f"Ошибка при парсинге предложения: {str(e)}")
    
    def create_offers_summary(self, offers, buyer_name):
        """Создание сводки предложений для заказчика"""
        if not offers:
            return "Нет предложений по вашей заявке."
        
        summary = f"📊 Сводка предложений для {buyer_name}\n"
        summary += f"⏰ Время: {datetime.now(self.timezone).strftime('%d.%m.%Y %H:%M')}\n"
        summary += "=" * 50 + "\n\n"
        
        for i, offer in enumerate(offers, 1):
            excel_info = f"\n📄 Excel файл: {offer['excel_filename']}" if offer.get('excel_filename') else ""
            summary += f"💼 **Предложение #{offer['id']}**\n"
            summary += f"👤 Поставщик: {offer['full_name']}\n"
            summary += f"📞 Телефон: {offer['phone_number']}\n"
            summary += f"💵 Общая сумма: {offer['total_amount']:,} сум\n"
            summary += f"📅 Дата: {offer['created_at'].strftime('%d.%m.%Y %H:%M')}{excel_info}\n"
            summary += "─" * 30 + "\n"
            
            # Детали товаров
            for j, item in enumerate(offer['items'], 1):
                summary += f"  {j}. {item['product_name']}\n"
                summary += f"     📊 Количество: {item['quantity']} {item['unit']}\n"
                summary += f"     💰 Цена за единицу: {item['price_per_unit']:,} сум\n"
                summary += f"     💵 Сумма: {item['total_price']:,} сум\n"
                if item['material_description']:
                    summary += f"     📝 Описание: {item['material_description']}\n"
                summary += "\n"
            
            summary += "─" * 30 + "\n\n"
        
        return summary
    
    def create_offers_excel(self, offers, buyer_name):
        """Создание Excel файла с предложениями для заказчика"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Предложения"
        
        # Заголовки
        headers = ['Предложение ID', 'Поставщик', 'Телефон', 'Товар', 'Количество', 'Единица', 
                  'Цена за единицу', 'Сумма', 'Описание', 'Общая сумма', 'Дата предложения']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Заполняем данными
        row = 2
        for offer in offers:
            for item in offer['items']:
                ws.cell(row=row, column=1, value=offer['id'])
                ws.cell(row=row, column=2, value=offer['full_name'])
                ws.cell(row=row, column=3, value=offer['phone_number'])
                ws.cell(row=row, column=4, value=item['product_name'])
                ws.cell(row=row, column=5, value=item['quantity'])
                ws.cell(row=row, column=6, value=item['unit'])
                ws.cell(row=row, column=7, value=item['price_per_unit'])
                ws.cell(row=row, column=8, value=item['total_price'])
                ws.cell(row=row, column=9, value=item['material_description'])
                ws.cell(row=row, column=10, value=offer['total_amount'])
                ws.cell(row=row, column=11, value=offer['created_at'].strftime('%d.%m.%Y %H:%M'))
                row += 1
        
        # Настройка ширины колонок
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def create_active_requests_excel(self, requests, seller_name):
        """Создание Excel файла с активными заявками для поставщиков"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Активные заявки"
        
        # Заголовки
        headers = ['Заявка ID', 'Заказчик', 'Поставщик', 'Объект', 'Товар', 'Количество', 
                  'Единица', 'Описание', 'Дата заявки']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Заполняем данными
        row = 2
        for request in requests:
            for item in request['items']:
                ws.cell(row=row, column=1, value=request['id'])
                ws.cell(row=row, column=2, value=request['buyer_name'])
                ws.cell(row=row, column=3, value=request['supplier_name'])
                ws.cell(row=row, column=4, value=request['object_name'])
                ws.cell(row=row, column=5, value=item['product_name'])
                ws.cell(row=row, column=6, value=item['quantity'])
                ws.cell(row=row, column=7, value=item['unit'])
                ws.cell(row=row, column=8, value=item['material_description'])
                ws.cell(row=row, column=9, value=request['created_at'].strftime('%d.%m.%Y %H:%M'))
                row += 1
        
        # Настройка ширины колонок
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def validate_excel_structure(self, file_content, file_type='request'):
        """Проверка структуры Excel файла"""
        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0)
            
            if file_type == 'request':
                required_columns = ['Обект номи', 'Махсулот номи', 'Миқдори', 'Ўлчов бирлиги', 'Материал изох']
            else:  # offer
                required_columns = ['Махсулот номи', 'Миқдори', 'Ўлчов бирлиги', 'Материал изох', 'нархи', 'Суммаси']
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return False, f"Отсутствуют обязательные колонки: {', '.join(missing_columns)}"
            
            return True, "Структура файла корректна"
            
        except Exception as e:
            return False, f"Ошибка при проверке файла: {str(e)}" 